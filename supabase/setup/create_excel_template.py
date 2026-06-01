import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

thin = Side(style='thin', color='CCCCCC')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── Sheet 1: Ptice - Import ─────────────────────────────────
ws = wb.active
ws.title = 'Ptice - Import'

DARK_BLUE   = '1E3A5F'
MID_BLUE    = '2D6A9F'
LIGHT_BLUE  = '4A90D9'
VLIGHT_BLUE = 'EBF3FB'
DARKEST     = '0D2137'
WHITE       = 'FFFFFF'
STRIPE      = 'F7FBFF'

columns = [
    ('oznaka',             14, 'OZNAKA *',        True,  'P001'),
    ('naziv',              18, 'NAZIV',           False, 'Zuti sampion'),
    ('spol',               8,  'SPOL *',          True,  'M'),
    ('vrsta',              22, 'VRSTA *',         True,  'Kanarinac stasa'),
    ('prstena_oznaka',     18, 'PRSTEN OZNAKA',   False, '2022-BA-15'),
    ('prsten_redni_broj',  12, 'PRSTEN BR.',      False, 15),
    ('godina',             10, 'GODINA',          False, 2022),
    ('datum_rodjenja',     14, 'DATUM RODJENJA',  False, '2022-03-10'),
    ('boja',               16, 'BOJA',            False, 'Zuta'),
    ('status_evidencije',  20, 'STATUS',          False, 'aktivna'),
    ('otac_oznaka',        14, 'OTAC OZNAKA',     False, ''),
    ('majka_oznaka',       14, 'MAJKA OZNAKA',    False, ''),
    ('napomena',           28, 'NAPOMENA',        False, ''),
]

# Red 1: Naslov
ws.merge_cells('A1:M1')
ws['A1'] = 'HatchPlan - Import ptica'
ws['A1'].font = Font(bold=True, color=WHITE, size=13)
ws['A1'].fill = PatternFill('solid', fgColor=DARKEST)
ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 28

# Red 2: razmak
ws.row_dimensions[2].height = 5

# Red 3: Header kolona
ws.row_dimensions[3].height = 36
opisi = [
    'Jedinstven ID u fajlu (npr. P001)',
    'Ime ptice (moze biti prazno)',
    'M / Z / ?',
    'Tacno kao u listi na sheetu Vrste ptica',
    'Oznaka prstena',
    'Cijeli broj',
    'Godina uzgoja (2022...)',
    'Format: YYYY-MM-DD',
    'Slobodan tekst',
    'aktivna/mlada/vanjska/uginula/prodata/poklonjena/ostalo',
    'Oznaka oca iz kolone A',
    'Oznaka majke iz kolone A',
    'Slobodan tekst',
]

for col_idx, (field, width, label, required, _) in enumerate(columns, 1):
    cell = ws.cell(row=3, column=col_idx, value=label)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill('solid', fgColor=MID_BLUE if required else LIGHT_BLUE)
    cell.alignment = center
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Red 4: Opisi
ws.row_dimensions[4].height = 30
for col_idx, opis in enumerate(opisi, 1):
    cell = ws.cell(row=4, column=col_idx, value=opis)
    cell.font = Font(color='888888', size=8, italic=True)
    cell.fill = PatternFill('solid', fgColor='F0F4F8')
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    cell.border = thin_border

# Redovi 5-9: Primjeri
examples = [
    ('P001', 'Zuti sampion', 'M', 'Kanarinac stasa', '2022-BA-15', 15, 2022, '2022-03-10', 'Zuta', 'aktivna', '', '', ''),
    ('P002', 'Bijelica', 'Z', 'Kanarinac stasa', '2022-BA-22', 22, 2022, '2022-03-15', 'Bijela', 'aktivna', '', '', ''),
    ('P003', 'Zlatko', 'M', 'Kanarinac stasa', '2023-BA-05', 5, 2023, '2023-04-01', 'Zuto-bijela', 'aktivna', 'P001', 'P002', 'Sin P001 i P002'),
    ('P004', '', 'Z', 'Stiaglic', 'XY-2021-08', 8, 2021, '', '', 'vanjska', '', '', 'Ptica od Ivana Kovaca - rodovnik'),
    ('P005', 'Mladi zuti', 'M', 'Kanarinac stasa', '2024-BA-11', 11, 2024, '2024-04-20', 'Zuta', 'mlada', 'P001', 'P002', ''),
]

for row_offset, example in enumerate(examples):
    row = 5 + row_offset
    ws.row_dimensions[row].height = 20
    fill_color = VLIGHT_BLUE if row_offset % 2 == 0 else STRIPE
    for col_idx, val in enumerate(example, 1):
        cell = ws.cell(row=row, column=col_idx, value=val if val != '' else None)
        cell.font = Font(color='555555', size=9, italic=True)
        cell.fill = PatternFill('solid', fgColor=fill_color)
        cell.alignment = left
        cell.border = thin_border

# Redovi 10-110: prazni za unos
for row in range(10, 111):
    ws.row_dimensions[row].height = 20
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=row, column=col_idx, value=None)
        cell.fill = PatternFill('solid', fgColor=WHITE)
        cell.font = Font(size=10)
        cell.alignment = left
        cell.border = thin_border

ws.freeze_panes = 'A5'

# ── Sheet 2: Vrste ptica ────────────────────────────────────
ws2 = wb.create_sheet('Vrste ptica')
ws2['A1'] = 'NAZIV VRSTE (kopiraj tacno u kolonu VRSTA)'
ws2['A1'].font = Font(bold=True, color=WHITE, size=10)
ws2['A1'].fill = PatternFill('solid', fgColor=DARK_BLUE)
ws2['A1'].alignment = center
ws2.column_dimensions['A'].width = 45
ws2.row_dimensions[1].height = 28

vrste = [
    'Kanarinac stasa',
    'Kanarinac boje',
    'Kanarinac pjevac',
    'Stiglic',
    'Zeba',
    'Zelenac',
]

for i, vrsta in enumerate(vrste, 2):
    cell = ws2.cell(row=i, column=1, value=vrsta)
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = thin_border
    cell.fill = PatternFill('solid', fgColor=VLIGHT_BLUE if i % 2 == 0 else WHITE)
    ws2.row_dimensions[i].height = 20

note = ws2.cell(row=len(vrste) + 3, column=1,
    value='Ako tvoja vrsta nije ovdje, provjeri Admin panel u aplikaciji i dodaj je.')
note.font = Font(size=9, italic=True, color='AA6600')
note.fill = PatternFill('solid', fgColor='FFF3CD')
note.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws2.row_dimensions[len(vrste) + 3].height = 30

# ── Sheet 3: Upute ─────────────────────────────────────────
ws3 = wb.create_sheet('Upute')
ws3.column_dimensions['A'].width = 70

upute = [
    ('HatchPlan - Upute za Excel import', True, DARKEST, WHITE, 14),
    ('', False, WHITE, '000000', 10),
    ('OBAVEZNA POLJA (oznacena tamnije u headeru)', True, DARK_BLUE, WHITE, 11),
    ('  OZNAKA - jedinstven identifikator u ovom fajlu (npr. P001, Zuti-2022)', False, VLIGHT_BLUE, '000000', 10),
    ('  SPOL - upisite M (muzjak), Z (zenka) ili ? (nepoznat)', False, VLIGHT_BLUE, '000000', 10),
    ('  VRSTA - mora biti TACNO kao naziv na sheetu "Vrste ptica"', False, VLIGHT_BLUE, '000000', 10),
    ('', False, WHITE, '000000', 10),
    ('RODITELJI (OTAC OZNAKA / MAJKA OZNAKA)', True, DARK_BLUE, WHITE, 11),
    ('  Upisite OZNAKU roditelja koja se nalazi u koloni A ovog fajla', False, VLIGHT_BLUE, '000000', 10),
    ('  Roditelj mora biti prisutan u istom fajlu', False, VLIGHT_BLUE, '000000', 10),
    ('  Ako roditelje ne znate ili ih nemate - ostavite prazno', False, VLIGHT_BLUE, '000000', 10),
    ('', False, WHITE, '000000', 10),
    ('STATUS EVIDENCIJE', True, DARK_BLUE, WHITE, 11),
    ('  aktivna    - vasa ptica, moze biti u parovima (DEFAULT)', False, VLIGHT_BLUE, '000000', 10),
    ('  mlada      - vasa mlada ptica, jos nije za parove', False, VLIGHT_BLUE, '000000', 10),
    ('  vanjska    - tudja ptica (samo za rodovnik, nije za parove)', False, VLIGHT_BLUE, '000000', 10),
    ('  uginula    - uginula ptica', False, VLIGHT_BLUE, '000000', 10),
    ('  prodata    - prodata ptica', False, VLIGHT_BLUE, '000000', 10),
    ('  poklonjena - poklonjena ptica', False, VLIGHT_BLUE, '000000', 10),
    ('', False, WHITE, '000000', 10),
    ('DATUM RODJENJA', True, DARK_BLUE, WHITE, 11),
    ('  Format mora biti: YYYY-MM-DD (npr. 2022-03-15)', False, VLIGHT_BLUE, '000000', 10),
    ('  Ako ne znate datum, ostavite prazno i upisite samo GODINU', False, VLIGHT_BLUE, '000000', 10),
    ('', False, WHITE, '000000', 10),
    ('VAZNA NAPOMENA', True, '0D6659', WHITE, 11),
    ('  Primjeri u redovima 5-9 su PRIMJERI - obrisite ih!', False, 'FFF3CD', '7A4F00', 10),
    ('  Posaljite popunjeni fajl administratoru koji ce obaviti import.', False, 'FFF3CD', '7A4F00', 10),
    ('  Sve ce biti provjereno prije unosa u bazu.', False, 'FFF3CD', '7A4F00', 10),
]

for i, (tekst, bold, bg, fg, size) in enumerate(upute, 1):
    cell = ws3.cell(row=i, column=1, value=tekst)
    cell.font = Font(bold=bold, size=size, color=fg)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws3.row_dimensions[i].height = 22 if tekst else 8

output_path = 'C:/Users/almirj/Desktop/SVELTE/uzgoj_ptica-app/supabase/setup/HatchPlan_Import_Ptica.xlsx'
wb.save(output_path)
print('Saved:', output_path)
